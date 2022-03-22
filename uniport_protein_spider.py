from bs4 import BeautifulSoup
import requests
import re
from queue import Queue
from openpyxl import Workbook
"""
giving page_text ，find and return protein name,Gene,Function,Path,PTM and seq
"""
class uni_pider(object):
    def __init__(self) -> None:
        retrieve = input('输入检索:')
        self.main_url = 'https://www.uniprot.org/uniprot/?query={}'.format(retrieve)+'&sort=score'
        r = requests.get(self.main_url)
        soup = BeautifulSoup(r.text)
        s = 'https://www.uniprot.org'
        s += soup.body.main.find(name='a',attrs={'class':'nextPageLink'})['href']
        print('总数量:' + str(soup.body.main.find(name='div',attrs={'class':'main-aside'}).find(name='strong',attrs={'class':'queryResultCount'}).get_text()))
        self.url = s.replace('offset=25','offset={}')
        self.main()


    def process(self,name):
        self.gene = ''
        self.seq = ''
        self.fun = []
        self.path = []
        self.ptm = []
        self.url = "https://www.uniprot.org/uniprot/"+name
        r = requests.get(self.url)
        text = r.text
        self.soup = BeautifulSoup(text)
        print("正在读取{}\n".format(name))
        self.get_text()
        self.get_Gene()
        self.get_seq()
        name_li = []
        name_li.append(name)
        self.ws_protein_name.append(name_li)
        gene_li = []
        gene_li.append(self.gene)
        self.ws_Gene.append(gene_li)
        self.ws_Function.append(self.fun)
        self.ws_Path.append(self.path)
        self.ws_PTM.append(self.ptm)
        seq_li = []
        seq_li.append(self.seq)
        self.ws_seq.append(seq_li)
        
        
                
    def get_Gene(self):
        try:
            self.gene = self.root.find(name='div',attrs={'id':'entry-overview'}).find(name='div',attrs={'id':'content-gene'}).h2.string
        except:
            self.gene = ''
            print('wu gene')
 
    def get_seq(self):
        try:
            seq_list = self.soup.body.main.find(name='div',attrs={'class':'main-aside'}).find(name='div',attrs={'id':'resultActionButtons'}).span.pre.string.split('\n')
        except:
            seq_list = ''
        for i in seq_list[1:]:
            self.seq += i
 
    def process_3table(self,text):
        rows_list = re.findall(r'\[[0-9-,\s]+\]&amp;key=.+?featureImage', text)
        pos = ''
        des = ''
        key = ''
        rows = []
        for row in rows_list:
            try:
                pos = re.findall(r'\[[0-9-,]+\]',row)[0][1:-1]
                pos = pos.replace(',','↔')
            except:
                pos = ''
            try: 
                key = re.findall(r'key=[A-Za-z\s]+',row)[0][4:]
            except:
                key = ''
            try:
                des = re.findall(r'\"text\".+?<\/span',row)[0][7:-6]
                des = re.sub(r'<.+>','',des)
            except:
                des = ''
            rows.append(key)
            rows.append(pos)
            rows.append(des)
        return rows
 
    def process_2table(self,text):
        rows_list = re.findall(r'\[[0-9-,\s]+\]&amp;key=[\s\S]+?featureImage', text)
        pos = ''
        key = ''
        rows = []
        for row in rows_list:
            try:
                pos = re.findall(r'\[[0-9-,]+\]',row)[0][1:-1]
                pos = pos.replace(',','↔')
            except:
                pos = ''
            try: 
                key = re.findall(r'key=[A-Za-z\s]+',row)[0][4:]
            except:
                key = ''
            rows.append(key)
            rows.append(pos)
        return rows
 
    def get_text(self):
        self.root = self.soup.body.main.find(name='div',attrs={'class':'main-aside'}).find(name='div',attrs={'class':'content entry_view_content up_entry swissprot'})
        if self.root ==None:
            self.root = self.soup.body.main.find(name='div',attrs={'class':'main-aside'}).find(name='div',attrs={'class':'content entry_view_content up_entry trembl'})
        self.func_text = str(self.root.find(name='div',attrs={'id':'function'}))
        self.path_text =str(self.root.find(name='div',attrs={'id':'pathology_and_biotech'}))
        self.ptm_text = str(self.root.find(name='div',attrs={'id':'ptm_processing'}))
        self.fun = self.process_2table(self.func_text)
        self.path = self.process_2table(self.path_text)
        self.ptm = self.process_3table(self.ptm_text)
 
 
        
    def op(self):
        self.wb = Workbook()
        self.wb.remove(self.wb['Sheet'])
        self.wb.create_sheet('protein name')
        self.wb.create_sheet('Gene')
        self.wb.create_sheet('Function')
        self.wb.create_sheet('Pathology & Biotech')
        self.wb.create_sheet('PTM Processing')
        self.wb.create_sheet('sequences')
        self.ws_protein_name = self.wb['protein name']
        self.ws_Gene = self.wb['Gene']
        self.ws_Function = self.wb['Function']
        self.ws_Path = self.wb['Pathology & Biotech']
        self.ws_PTM = self.wb['PTM Processing']
        self.ws_seq = self.wb['sequences']
        


    def get_name(self):
        self.name_queue = Queue()
        for i in range(0,self.end_num,25):
            current_url = self.url.format(i)
            r = requests.get(current_url)
            name_soup = BeautifulSoup(r.text)
            n = name_soup.body.main.find(name='div',attrs={'class':'main-aside'}).find(name='div',attrs={'class':'content results'}).find(name='div',attrs={'id':'resultsArea'}).form.table.tbody.find_all('tr')
            for i in n:
                self.name_queue.put(i['id'])
    
    def main(self):
        self.end_num = int(input('爬取数量:'))
        filename = input('保存文件名:')
        self.op()
        self.get_name()
        for i in range(self.end_num):
            na = self.name_queue.get()
            self.process(na)
        self.wb.save(filename=filename+'.xlsx')



if __name__ == '__main__':    
    a = uni_pider()

